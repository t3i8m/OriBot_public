import pandas
from openpyxl import load_workbook
import os

user_chat_id_collector = []
users_db_json={"Users":[{"Username": "oribot","Name":"Oribot","Chat Id":"1234","Language":"en" ,"Time of registartion":"05.11.2022","Premium status":True,"Change language":False, "Location":"Riga", "Lat": 56.95, "Long":24.1, "Premium expiring date":"27.04.2023", "Promo code expiring date": "20.04.2023"}]}
counter = 0

def appending_new_user(user, users_db_json=""):
    if users_db_json!="":
        try:
            users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": user.promo_code_expiring_date[0]})
        except Exception as ex:
            print(ex)
            users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False})

        try:
            if user.premium_expiring_date==None and user.promo_code_expiring_date==None:
                users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False})
            elif user.premium_expiring_date!=None and user.promo_code_expiring_date==None:
                users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": False})
            elif user.premium_expiring_date==None and user.promo_code_expiring_date!=None:
                users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": user.promo_code_expiring_date[0]})
            elif user.premium_expiring_date!=None and user.promo_code_expiring_date!=None:
                users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": user.promo_code_expiring_date[0]})
        except Exception:
                users_db_json["Users"].append({"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False})

    for root, dir, files in os.walk(fr"\oribot_main\releated_files"):
        if "Oribot_users.xlsx" not in files:
            print(users_db_json, user_chat_id_collector)
            user_chat_id_collector.append(str(user.chat_id))
            create_xlsx(users_db_json)
            return
        else:
            update_user(user)
        break
    return

def update_user(user):
    try:
        book = load_workbook(fr"\oribot_main\releated_files\Oribot_users.xlsx")
    except Exception as ex:
        book = load_workbook(fr"\oribot_support\releated_files_support\Oribot_users_support.xlsx")
    # try:
    #     updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": user.promo_code_expiring_date[0]}
    # except Exception as ex:

    #     print(ex)
    #     updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False}
    try:
        if user.premium_expiring_date==None and user.promo_code_expiring_date==None:
            updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False}
        elif user.premium_expiring_date!=None and user.promo_code_expiring_date==None:
            updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": False}
        elif user.premium_expiring_date==None and user.promo_code_expiring_date!=None:
            updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": user.promo_code_expiring_date[0]}
        elif user.premium_expiring_date!=None and user.promo_code_expiring_date!=None:
            updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":user.premium_expiring_date[0], "Promo code expiring date": user.promo_code_expiring_date[0]}
    except Exception:
            updated_user = {"Username": user.user_name,"Name":user.user_first_name,"Chat Id":str(user.chat_id),"Language":user.lang, "Time of registartion":user.time_of_registration,"Premium status":user.premium,"Change language":user.change_lang,"Location":user.location,"Lat":user.lat, "Long":user.long, "Premium expiring date":False, "Promo code expiring date": False}

    # if str(user.chat_id) in set(user_chat_id_collector):
    #     for index,row in enumerate(book.active):
    #         for index_k, cell in enumerate(row):
    #             if cell.value == str(user.chat_id):
    #                 for dict_v,cell_k in zip(updated_user.values(),row):
    #                     cell_k.value = dict_v
    #                 break
    # else:
    #     book.active.append(list(updated_user.values())) 
    #     user_chat_id_collector.append(str(user.chat_id))
    sheet = book.active

    # Найти индекс столбца с заголовком "Chat Id"
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    id_column_index = None
    for index, cell in enumerate(header_row):
        if cell.value == "Chat Id":
            id_column_index = index + 1
            break

    # Получить значения из столбца "Chat Id"
    id_collector_from_excel_file = []
    for row in sheet.iter_rows(min_row=2, max_col=id_column_index, values_only=True):
        id_collector_from_excel_file.append(str(row[id_column_index-1]))

    if str(user.chat_id) in set(user_chat_id_collector) or str(user.chat_id) in set(id_collector_from_excel_file):
        for index,row in enumerate(book.active):
            for index_k, cell in enumerate(row):
                if cell.value == str(user.chat_id):
                    for dict_v,cell_k in zip(updated_user.values(),row):
                        cell_k.value = dict_v
                    break
        user_chat_id_collector.append(str(user.chat_id))
    else:
        book.active.append(list(updated_user.values())) 
        user_chat_id_collector.append(str(user.chat_id))
    book.save(fr"\oribot_main\releated_files\Oribot_users.xlsx")
    book.save(fr"\oribot_support\releated_files_support\Oribot_users_support.xlsx") 
    return True


def create_xlsx(users_db_json):
    df = pandas.DataFrame(users_db_json["Users"])
    df.to_excel(fr"\oribot_main\releated_files\Oribot_users.xlsx",sheet_name="Oribot_users", index=False)
    df.to_excel(fr"\oribot_support\releated_files_support\Oribot_users_support.xlsx",sheet_name="Oribot_users", index=False)
    return
    
